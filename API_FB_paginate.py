"""API_FB_paginate.py
The script prints out a randomlly selected comment from current comments related to the post. Than the comments are saved into xmlx file and the selected one is highlited.
It requires to enter globals - ACCESS_TOKEN, POST_ID or paste them into new file 'token_and_id.txt'. If you like to, edit PAGINATE_LIMIT."""

import facebook
import random
import openpyxl
from openpyxl.styles import Font
import os 


DIR_PATH = os.path.dirname(os.path.realpath(__file__))
FILENAME = "comments_winner.xlsx"
PAGINATE_LIMIT = 2000


def get_token_and_id(file):
    """"Works with file where the first line is token and the second line is id."""
    with open(file) as f:
        lines = f.read().splitlines()
        token, id = lines[0], lines[1]
    return token, id


ACCESS_TOKEN, POST_ID = get_token_and_id(DIR_PATH + "\\" + "token_and_id.txt")


def load_comments(token, post_id, set_limit=2000):
    graph = facebook.GraphAPI(access_token=token, version="3.1")
    page = 0
    while True:
        comments = graph.get_connections(id=post_id, connection_name="comments", limit=set_limit, offset=page*set_limit)
        if comments['data'] == []: break
        page += 1
        yield comments


def get_winning_comment(comments):
    comments_list = [comment for comment in comments['data'] if comment['message']]
    comment = random.choice(comments_list)
    comment_id, content = comment['id'], comment['message']

    return comment_id, content


def create_new_csv(filename):
    """Returns False when file already exists"""
    try:
        with open(filename, "x", newline=""): return True
    except FileExistsError:
        return False    


def add_row(worksheet, data_list, row=1):
    for i, cell in enumerate(data_list):
        worksheet.cell(row=row, column=i+1).value = cell


def upload_comments_to_xmlx(comments, winners_id):
    """Returns False when file has been previously created (for the reason not to overcreate a file and change the winner)."""
    if not create_new_csv(f"{DIR_PATH}\\{FILENAME}"):
        return False

    workbook = openpyxl.Workbook()
    worksheet = workbook.active


    # upload all comments to a sheet section:
    header_fields = ["WINNER", "TIME", "ID", "MESSAGE"]
    add_row(worksheet, header_fields)

    for i, comment in enumerate(comments['data']):
        comment_fields = ["", comment['created_time'], comment['id'], comment['message']]
        add_row(worksheet, comment_fields, i+2)

        if comment['id'] == winners_id:
            winners_row = i+2 # +1 for no '0' row and +1 for the header
            winners_comment = comment
            add_row(worksheet, ["->"], row=winners_row)
            for c in range(1, 5):
                worksheet.cell(row=winners_row, column=c).font = Font(color="00FF00")

    

    
    # upload the winners comment to second sheet in the file:    
    new_sheet_name = "winners_sheet"
    workbook.create_sheet(new_sheet_name)
    worksheet = workbook[new_sheet_name]
    

    winners_fields = ["WINNER:", winners_comment['created_time'], winners_comment['id'], winners_comment['message']]
    add_row(worksheet, header_fields)
    add_row(worksheet, winners_fields)


    workbook.save(f"{DIR_PATH}\\{FILENAME}")
    return True
    

if __name__ == '__main__':
    
    # paginate section
    all_comments = {"data": []}
    for comments in load_comments(ACCESS_TOKEN, POST_ID, set_limit=PAGINATE_LIMIT):
        # print("A page:", [x["message"] for x in comments['data'] if x['message']]) # Un-comment see each page separately
        all_comments["data"].extend(comments["data"])

    # print(all_comments)

    comment_id, comment_content = get_winning_comment(all_comments)

    # xmlx related section
    is_uploaded = upload_comments_to_xmlx(all_comments, comment_id)
    if is_uploaded:
        print(f"INFO:   Comments were sucessfully save to '{FILENAME}'.")
        print(f'RESULT: A comment id#{comment_id} with content "{comment_content}" was randomly selected.')
    else:
        print(f"INFO:   ! No success with file and/or comments upload. (delete '{FILENAME}' or change its name)")